import os
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Optional, Dict, Any

from backend.database import get_db_connection, XML_STORAGE_DIR, get_nfe_detail
from backend.services.danfe_service import generate_danfe_pdf, build_synthetic_nfe_xml


def generate_pacote_contabil_zip(
    ano: int,
    mes: int,
    empresa_cnpj: Optional[str] = None,
    incluir_pdfs: bool = True
) -> io.BytesIO:
    """
    Gera um arquivo ZIP estruturado contendo:
    - 01-Entradas/ (XMLs de Compras)
    - 02-Saidas/ (XMLs de Vendas)
    - 03-Canceladas_e_Eventos/
    - 04-DANFEs_PDF/ (Se incluir_pdfs=True)
    - Relatorio_Fiscal_Fechamento_{mes:02d}_{ano}.xlsx
    """
    clean_cnpj = "".join(c for c in str(empresa_cnpj) if c.isdigit()) if empresa_cnpj else None
    mes_prefix = f"{ano}-{mes:02d}"

    with get_db_connection() as conn:
        cursor = conn.cursor()

        # 1. Notas de Entrada (Compras)
        q_ent = "SELECT * FROM nfe_docs WHERE (tipo_doc = 0 OR tipo_doc IS NULL) AND substr(data_emissao, 1, 7) = ?"
        p_ent = [mes_prefix]
        if clean_cnpj:
            q_ent += " AND (empresa_cnpj = ? OR destinatario_cnpj LIKE ?)"
            p_ent.extend([clean_cnpj, f"%{clean_cnpj}%"])
        cursor.execute(q_ent, p_ent)
        entradas = [dict(r) for r in cursor.fetchall()]

        # 2. Notas de Saída (Vendas)
        q_sai = "SELECT * FROM nfe_docs WHERE tipo_doc = 1 AND substr(data_emissao, 1, 7) = ?"
        p_sai = [mes_prefix]
        if clean_cnpj:
            q_sai += " AND (empresa_cnpj = ? OR emitente_cnpj LIKE ?)"
            p_sai.extend([clean_cnpj, f"%{clean_cnpj}%"])
        cursor.execute(q_sai, p_sai)
        saidas = [dict(r) for r in cursor.fetchall()]

        # 3. Eventos
        cursor.execute("""
            SELECT e.*, d.numero, d.tipo_doc
            FROM nfe_events e
            JOIN nfe_docs d ON e.chave = d.chave
            WHERE substr(e.dh_evento, 1, 7) = ?
        """, (mes_prefix,))
        eventos = [dict(r) for r in cursor.fetchall()]

    # Cria o arquivo ZIP em memória
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Adiciona XMLs de Entradas
        for doc in entradas:
            chave = doc["chave"]
            xml_bytes = _get_or_build_xml(doc)
            if xml_bytes:
                zf.writestr(f"01-Entradas/{chave}.xml", xml_bytes)
                if incluir_pdfs:
                    pdf_buf = generate_danfe_pdf(xml_bytes)
                    if pdf_buf:
                        zf.writestr(f"04-DANFEs_PDF/DANFE_ENTRADA_{doc.get('numero', '')}_{chave}.pdf", pdf_buf.getvalue())

        # Adiciona XMLs de Saídas
        for doc in saidas:
            chave = doc["chave"]
            xml_bytes = _get_or_build_xml(doc)
            if xml_bytes:
                pasta = "03-Canceladas_e_Eventos" if "cancelad" in (doc.get("situacao") or "").lower() else "02-Saidas"
                zf.writestr(f"{pasta}/{chave}.xml", xml_bytes)
                if incluir_pdfs:
                    pdf_buf = generate_danfe_pdf(xml_bytes)
                    if pdf_buf:
                        zf.writestr(f"04-DANFEs_PDF/DANFE_SAIDA_{doc.get('numero', '')}_{chave}.pdf", pdf_buf.getvalue())

        # Adiciona Relatório Excel Oficial
        excel_bytes = _generate_excel_fechamento(ano, mes, entradas, saidas, empresa_cnpj)
        zf.writestr(f"Relatorio_Fiscal_Fechamento_{mes:02d}_{ano}.xlsx", excel_bytes)

    zip_buf.seek(0)
    return zip_buf


def _get_or_build_xml(doc: Dict[str, Any]) -> bytes:
    chave = doc["chave"]
    disk_path = os.path.join(XML_STORAGE_DIR, f"{chave}.xml")
    if os.path.exists(disk_path):
        try:
            with open(disk_path, "rb") as f:
                c = f.read()
                if len(c) > 50:
                    return c
        except Exception:
            pass

    if doc.get("xml_raw") and len(doc["xml_raw"]) > 50:
        return doc["xml_raw"].encode("utf-8")

    full_doc = get_nfe_detail(chave) or doc
    return build_synthetic_nfe_xml(full_doc)


def _generate_excel_fechamento(ano: int, mes: int, entradas: list, saidas: list, empresa_cnpj: Optional[str]) -> bytes:
    wb = openpyxl.Workbook()

    # Cores e Estilos
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    sub_fill = PatternFill(start_color="2874A6", end_color="2874A6", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True)
    normal_font = Font(name="Arial", size=10)

    # 1. ABA RESUMO GERAL
    ws_res = wb.active
    ws_res.title = "Resumo Geral do Mês"

    ws_res.cell(1, 1, f"FECHAMENTO FISCAL & CONTÁBIL - COMPETÊNCIA {mes:02d}/{ano}").font = Font(name="Arial", size=14, bold=True, color="1B4F72")
    ws_res.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Empresa: {empresa_cnpj or 'Todas as Filiais do Grupo'}").font = Font(name="Arial", size=10, italic=True)

    ws_res.append([])
    headers_res = ["Indicador Fiscal", "Qtd. Documentos", "Valor Total dos Produtos", "ICMS Destacado", "PIS", "COFINS"]
    ws_res.append(headers_res)
    for col in range(1, len(headers_res) + 1):
        cell = ws_res.cell(4, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    tot_sai_v = sum(float(d.get("valor_total") or 0.0) for d in saidas if "cancelad" not in (d.get("situacao") or "").lower())
    tot_sai_icms = sum(float(d.get("valor_icms") or 0.0) for d in saidas if "cancelad" not in (d.get("situacao") or "").lower())
    tot_ent_v = sum(float(d.get("valor_total") or 0.0) for d in entradas)
    tot_ent_icms = sum(float(d.get("valor_icms") or 0.0) for d in entradas)

    ws_res.append(["1. Faturamento Bruto (Saídas / Vendas)", len(saidas), tot_sai_v, tot_sai_icms, tot_sai_v * 0.0065, tot_sai_v * 0.03])
    ws_res.append(["2. Compras & Insumos (Entradas)", len(entradas), tot_ent_v, tot_ent_icms, tot_ent_v * 0.0065, tot_ent_v * 0.03])
    ws_res.append(["3. Saldo Operacional Líquido", len(saidas) + len(entradas), tot_sai_v - tot_ent_v, tot_sai_icms - tot_ent_icms, 0.0, 0.0])

    for row in range(5, 8):
        ws_res.cell(row, 1).font = bold_font
        for c in range(2, 7):
            cell = ws_res.cell(row, c)
            cell.font = normal_font
            if c >= 3:
                cell.number_format = "R$ #,##0.00"

    # 2. ABA SAÍDAS (VENDAS)
    ws_sai = wb.create_sheet(title="Notas de Saída (Vendas)")
    headers_sai = ["Data Emissão", "Número", "Série", "Chave de Acesso", "Cliente / Destinatário", "CPF / CNPJ", "UF", "Valor Total (R$)", "ICMS (R$)", "Situação"]
    ws_sai.append(headers_sai)
    for col in range(1, len(headers_sai) + 1):
        cell = ws_sai.cell(1, col)
        cell.fill = sub_fill
        cell.font = header_font

    for d in saidas:
        d_emi = (d.get("data_emissao") or "")[:10]
        ws_sai.append([
            d_emi,
            d.get("numero", ""),
            d.get("serie", "1"),
            d.get("chave", ""),
            d.get("destinatario_nome", ""),
            d.get("destinatario_cnpj", ""),
            d.get("destinatario_uf", "SP"),
            float(d.get("valor_total") or 0.0),
            float(d.get("valor_icms") or 0.0),
            d.get("situacao", "Autorizada")
        ])

    # 3. ABA ENTRADAS (COMPRAS)
    ws_ent = wb.create_sheet(title="Notas de Entrada (Compras)")
    headers_ent = ["Data Emissão", "Número", "Chave de Acesso", "Fornecedor / Emitente", "CNPJ Fornecedor", "UF", "Valor Total (R$)", "ICMS (R$)", "Situação"]
    ws_ent.append(headers_ent)
    for col in range(1, len(headers_ent) + 1):
        cell = ws_ent.cell(1, col)
        cell.fill = sub_fill
        cell.font = header_font

    for d in entradas:
        d_emi = (d.get("data_emissao") or "")[:10]
        ws_ent.append([
            d_emi,
            d.get("numero", ""),
            d.get("chave", ""),
            d.get("emitente_nome", ""),
            d.get("emitente_cnpj", ""),
            d.get("emitente_uf", "SP"),
            float(d.get("valor_total") or 0.0),
            float(d.get("valor_icms") or 0.0),
            d.get("situacao", "Autorizada")
        ])

    # Ajusta largura das colunas
    for ws in [ws_res, ws_sai, ws_ent]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
