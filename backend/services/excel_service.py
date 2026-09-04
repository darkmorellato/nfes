import io
from typing import Optional, Dict
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from backend.database import list_certificates_db, get_db_connection
from backend.utils.excel_helpers import (
    TOTAL_FILL, BOLD_FONT,
    CURRENCY_FMT, NUM_FMT, format_cnpj, apply_header_row, apply_title, apply_subtitle, auto_adjust_columns,
)


def generate_fiscal_excel(mes: Optional[int] = None, ano: Optional[int] = None) -> io.BytesIO:
    """Gera uma planilha Excel .xlsx profissional e estilizada com abas separadas por empresa."""
    now = datetime.now()
    ano = ano or now.year
    mes = mes or now.month
    mes_str = f"{ano:04d}-{mes:02d}"

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    certs = list_certificates_db()

    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT chave, empresa_cnpj, numero, serie, modelo, emitente_cnpj, emitente_nome,
                   destinatario_cnpj, destinatario_nome, data_emissao,
                   valor_total, valor_icms, valor_pis, valor_cofins, valor_ipi, situacao
            FROM nfe_docs
            WHERE data_emissao LIKE ?
            ORDER BY empresa_cnpj, data_emissao ASC
        """, (f"{mes_str}%",))
        docs_mes = [dict(r) for r in cursor.fetchall()]

    # ================================================================
    # 1. ABA RESUMO CONSOLIDADO
    # ================================================================
    ws_resumo = wb.create_sheet(title="Resumo Consolidado")
    ws_resumo.views.sheetView[0].showGridLines = True

    apply_title(ws_resumo, 1, 1, f"FECHAMENTO FISCAL CONSOLIDADO - {mes:02d}/{ano:04d}")
    apply_subtitle(ws_resumo, 2, 1, f"Gerado em {now.strftime('%d/%m/%Y %H:%M:%S')} - Grupo Empresarial Multi-Empresas")

    headers_emp = ["CNPJ Empresa", "Razão Social", "Qtd Notas", "Total Compras (R$)", "Total ICMS (R$)", "Total PIS (R$)", "Total COFINS (R$)"]
    ws_resumo.append([])
    ws_resumo.append(["RESUMO POR EMPRESA"])
    ws_resumo["A4"].font = Font(name="Calibri", size=12, bold=True, color="2C3E50")

    ws_resumo.append(headers_emp)
    header_row_idx = 5
    apply_header_row(ws_resumo, header_row_idx, headers_emp)

    curr_row = 6
    for c in certs:
        c_cnpj = c["cnpj"]
        c_docs = [d for d in docs_mes if d.get("empresa_cnpj") == c_cnpj or c_cnpj in (d.get("destinatario_cnpj") or "")]
        qtd = len(c_docs)
        v_tot = sum(float(d.get("valor_total") or 0) for d in c_docs)
        v_icms = sum(float(d.get("valor_icms") or 0) for d in c_docs)
        v_pis = sum(float(d.get("valor_pis") or 0) for d in c_docs)
        v_cof = sum(float(d.get("valor_cofins") or 0) for d in c_docs)

        cnpj_fmt = format_cnpj(c_cnpj)

        ws_resumo.append([cnpj_fmt, c["razao_social"], qtd, v_tot, v_icms, v_pis, v_cof])

        ws_resumo.cell(row=curr_row, column=3).number_format = NUM_FMT
        for col_i in range(4, 8):
            ws_resumo.cell(row=curr_row, column=col_i).number_format = CURRENCY_FMT
        curr_row += 1

    total_row = curr_row
    ws_resumo.append([
        "TOTAL DO GRUPO", "", f"=SUM(C6:C{total_row-1})", f"=SUM(D6:D{total_row-1})",
        f"=SUM(E6:E{total_row-1})", f"=SUM(F6:F{total_row-1})", f"=SUM(G6:G{total_row-1})"
    ])
    for col_i in range(1, 8):
        c_cell = ws_resumo.cell(row=total_row, column=col_i)
        c_cell.fill = TOTAL_FILL
        c_cell.font = BOLD_FONT
        if col_i >= 4:
            c_cell.number_format = CURRENCY_FMT

    # ================================================================
    # 2. ABAS INDIVIDUAIS POR EMPRESA
    # ================================================================
    headers_doc = [
        "Data Emissão", "Número", "Série", "Chave de Acesso", "CNPJ Emitente",
        "Razão Social Emitente", "Valor Total (R$)", "ICMS (R$)", "PIS (R$)", "COFINS (R$)", "IPI (R$)", "Situação"
    ]

    for c in certs:
        c_cnpj = c["cnpj"]
        short_name = (c["razao_social"][:25]).strip()
        for ch in ["\\", "/", "?", "*", "[", "]", ":"]:
            short_name = short_name.replace(ch, "_")

        ws = wb.create_sheet(title=short_name)
        ws.views.sheetView[0].showGridLines = True

        apply_title(ws, 1, 1, f"{c['razao_social']} - CNPJ {c['cnpj']}")
        apply_subtitle(ws, 2, 1, f"Fechamento Fiscal de Compras e Entradas: {mes:02d}/{ano:04d}")

        ws.append([])
        ws.append(headers_doc)
        h_row = 4
        apply_header_row(ws, h_row, headers_doc)

        c_docs = [d for d in docs_mes if d.get("empresa_cnpj") == c_cnpj or c_cnpj in (d.get("destinatario_cnpj") or "")]

        r_idx = 5
        for d in c_docs:
            dt_str = d.get("data_emissao", "")
            if dt_str and len(dt_str) >= 10:
                try:
                    dt_val = datetime.fromisoformat(dt_str.replace("Z", "")).strftime("%d/%m/%Y")
                except Exception:
                    dt_val = dt_str[:10]
            else:
                dt_val = dt_str

            ws.append([
                dt_val,
                d.get("numero", ""),
                d.get("serie", "1"),
                d.get("chave", ""),
                d.get("emitente_cnpj", ""),
                d.get("emitente_nome", ""),
                float(d.get("valor_total") or 0),
                float(d.get("valor_icms") or 0),
                float(d.get("valor_pis") or 0),
                float(d.get("valor_cofins") or 0),
                float(d.get("valor_ipi") or 0),
                d.get("situacao", "Autorizada"),
            ])

            for col_i in range(7, 12):
                ws.cell(row=r_idx, column=col_i).number_format = CURRENCY_FMT
            r_idx += 1

        if c_docs:
            tot_r = r_idx
            ws.append([
                "TOTAL", "", "", f"{len(c_docs)} nota(s)", "", "",
                f"=SUM(G5:G{tot_r-1})", f"=SUM(H5:H{tot_r-1})", f"=SUM(I5:I{tot_r-1})",
                f"=SUM(J5:J{tot_r-1})", f"=SUM(K5:K{tot_r-1})", ""
            ])
            for col_i in range(1, 13):
                c_cell = ws.cell(row=tot_r, column=col_i)
                c_cell.fill = TOTAL_FILL
                c_cell.font = BOLD_FONT
                if 7 <= col_i <= 11:
                    c_cell.number_format = CURRENCY_FMT

    auto_adjust_columns(ws_resumo)
    for sheet in wb.worksheets:
        if sheet != ws_resumo:
            auto_adjust_columns(sheet)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _safe_sheet_title(name: str, suffix: str = "") -> str:
    """Remove caracteres ilegais e limita a 31 chars (limite do Excel)."""
    s = (name or "Empresa").strip()
    for ch in ["\\", "/", "?", "*", "[", "]", ":"]:
        s = s.replace(ch, "_")
    full = f"{s}{suffix}"
    return full[:31] if len(full) > 31 else full


def generate_fiscal_excel_detalhado(mes: int, ano: int) -> io.BytesIO:
    """Gera planilha Excel detalhada para a contabilidade.

    Estrutura:
      - Aba ``Resumo`` com totais gerais por empresa e tipo (entrada/saída).
      - Uma aba por empresa com duas seções separadas:
          * ``SAÍDA`` (notas emitidas por nós)
          * ``ENTRADA`` (notas de fornecedores)
      Cada seção lista os itens produto por produto com impostos.
    """
    now = datetime.now()
    mes_str = f"{ano:04d}-{mes:02d}"

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    certs = list_certificates_db()

    with get_db_connection() as conn:
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT chave, empresa_cnpj, tipo_doc, numero, serie, modelo,
                   emitente_cnpj, emitente_nome, destinatario_cnpj, destinatario_nome,
                   data_emissao, valor_total, valor_icms, valor_pis, valor_cofins, valor_ipi,
                   situacao
            FROM nfe_docs
            WHERE data_emissao LIKE ?
            ORDER BY empresa_cnpj, tipo_doc ASC, data_emissao ASC, numero ASC
            """,
            (f"{mes_str}%",),
        )
        docs_mes = [dict(r) for r in cursor.fetchall()]

        chaves = [d["chave"] for d in docs_mes if d.get("chave")]
        items_por_chave: Dict[str, list] = {ch: [] for ch in chaves}

        xml_por_chave: Dict[str, str] = {}

        if chaves:
            placeholders = ",".join("?" * len(chaves))
            cursor.execute(
                f"""
                SELECT chave, n_item, codigo, ean, descricao, ncm, cfop, unidade,
                       quantidade, valor_unitario, valor_total, cst, v_icms
                FROM nfe_items
                WHERE chave IN ({placeholders})
                ORDER BY chave, n_item ASC
                """,
                tuple(chaves),
            )
            for it in cursor.fetchall():
                items_por_chave.setdefault(it["chave"], []).append(dict(it))

            cursor.execute(
                f"SELECT chave, xml_raw FROM nfe_docs WHERE chave IN ({placeholders}) AND (xml_raw IS NOT NULL AND xml_raw != '')",
                tuple(chaves),
            )
            for r in cursor.fetchall():
                xml_por_chave[r["chave"]] = r["xml_raw"] or ""

    # FALLBACK: extrai itens do xml_raw
    try:
        from backend.services.danfe_service import parse_nfe_xml
    except Exception:
        parse_nfe_xml = None

    chaves_sem_itens = [ch for ch, its in items_por_chave.items() if not its and xml_por_chave.get(ch)]
    for ch in chaves_sem_itens:
        xml_raw = xml_por_chave.get(ch) or ""
        if not xml_raw or not parse_nfe_xml:
            continue
        try:
            parsed = parse_nfe_xml(xml_raw.encode("utf-8")) or {}
        except Exception:
            parsed = {}
        produtos = parsed.get("produtos") or []
        if not produtos:
            continue
        itens_norm = []
        for p in produtos:
            itens_norm.append({
                "n_item": p.get("n_item") or p.get("item") or "",
                "codigo": p.get("codigo") or "",
                "ean": p.get("ean") or "",
                "descricao": p.get("descricao") or "",
                "cfop": p.get("cfop") or "",
                "ncm": p.get("ncm") or "",
                "cst": p.get("cst") or "",
                "unidade": p.get("unidade") or "",
                "quantidade": float(p.get("quantidade") or 0),
                "valor_unitario": float(p.get("valor_unitario") or 0),
                "valor_total": float(p.get("valor_total") or 0),
                "v_icms": float(p.get("v_icms") or 0),
            })
        if itens_norm:
            items_por_chave[ch] = itens_norm

    tipo_label = {0: "ENTRADA", 1: "SAÍDA"}

    def fmt_data(val: str) -> str:
        if not val:
            return ""
        try:
            return datetime.fromisoformat(val.replace("Z", "")).strftime("%d/%m/%Y")
        except Exception:
            return str(val)[:10]

    def operacao_para(d: dict) -> str:
        return tipo_label.get(int(d.get("tipo_doc") or 0), "OUTRA")

    # ============================================================
    # ABA RESUMO
    # ============================================================
    ws_resumo = wb.create_sheet(title="Resumo")
    apply_title(ws_resumo, 1, 1, f"RELATÓRIO CONTÁBIL DETALHADO - {mes:02d}/{ano:04d}")
    apply_subtitle(ws_resumo, 2, 1, f"Gerado em {now.strftime('%d/%m/%Y %H:%M:%S')} - Grupo Multi-Empresas")
    ws_resumo.append([])

    headers_resumo = [
        "CNPJ Empresa", "Razão Social", "Operação",
        "Qtd NF-e", "Qtd Itens", "Valor Total (R$)",
        "ICMS (R$)", "PIS (R$)", "COFINS (R$)", "IPI (R$)",
    ]
    ws_resumo.append(headers_resumo)
    apply_header_row(ws_resumo, 4, headers_resumo)

    docs_por_empresa: Dict[str, list] = {}
    for d in docs_mes:
        docs_por_empresa.setdefault(d.get("empresa_cnpj") or "", []).append(d)

    row = 5
    for cert in certs:
        cnpj = cert.get("cnpj") or ""
        emp_docs = docs_por_empresa.get(cnpj, [])
        for op_nome in ("SAÍDA", "ENTRADA"):
            docs_op = [d for d in emp_docs if operacao_para(d) == op_nome]
            if not docs_op:
                continue
            qtd_itens = sum(len(items_por_chave.get(d["chave"], [])) for d in docs_op)
            ws_resumo.append([
                format_cnpj(cnpj),
                cert.get("razao_social") or cnpj,
                op_nome,
                len(docs_op),
                qtd_itens,
                sum(float(d.get("valor_total") or 0) for d in docs_op),
                sum(float(d.get("valor_icms") or 0) for d in docs_op),
                sum(float(d.get("valor_pis") or 0) for d in docs_op),
                sum(float(d.get("valor_cofins") or 0) for d in docs_op),
                sum(float(d.get("valor_ipi") or 0) for d in docs_op),
            ])
            for col_i in (6, 7, 8, 9, 10):
                ws_resumo.cell(row=row, column=col_i).number_format = CURRENCY_FMT
            row += 1

    # Linha TOTAL GERAL
    total_nfe = len(docs_mes)
    total_itens = sum(len(items_por_chave.get(d["chave"], [])) for d in docs_mes)
    ws_resumo.append([
        "TOTAL GERAL", "", "",
        total_nfe,
        total_itens,
        sum(float(d.get("valor_total") or 0) for d in docs_mes),
        sum(float(d.get("valor_icms") or 0) for d in docs_mes),
        sum(float(d.get("valor_pis") or 0) for d in docs_mes),
        sum(float(d.get("valor_cofins") or 0) for d in docs_mes),
        sum(float(d.get("valor_ipi") or 0) for d in docs_mes),
    ])
    for col_i in range(1, len(headers_resumo) + 1):
        cell = ws_resumo.cell(row=row, column=col_i)
        cell.fill = TOTAL_FILL
        cell.font = BOLD_FONT
        if col_i >= 6:
            cell.number_format = CURRENCY_FMT
    row += 1

    auto_adjust_columns(ws_resumo)

    # ============================================================
    # ABAS POR EMPRESA — seções SAÍDA e ENTRADA separadas
    # ============================================================
    headers_item = [
        "Nº NF", "Data Emissão", "Chave de Acesso",
        "Emitente (CNPJ)", "Emitente (Nome)",
        "Destinatário (CNPJ)", "Destinatário (Nome)",
        "Situação",
        "Item", "Cód. Produto", "EAN", "Descrição do Produto", "CFOP", "NCM", "CST",
        "Unidade", "Quantidade", "Valor Unit. (R$)", "Valor Total Item (R$)",
        "ICMS Item (R$)",
    ]

    SECTION_FILL = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    SECTION_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    ENTRADA_FILL = openpyxl.styles.PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    SAIDA_FILL = openpyxl.styles.PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    used_titles = set()
    for cert in certs:
        cnpj = cert.get("cnpj") or ""
        emp_docs = [d for d in docs_mes if d.get("empresa_cnpj") == cnpj]
        if not emp_docs:
            continue

        razao = cert.get("razao_social") or cnpj or "Empresa"
        sheet_title = _safe_sheet_title(razao[:20])
        original = sheet_title
        n = 2
        while sheet_title in used_titles:
            sheet_title = _safe_sheet_title(original, suffix=f"_{n}")
            n += 1
        used_titles.add(sheet_title)

        ws = wb.create_sheet(title=sheet_title)
        apply_title(ws, 1, 1, f"{razao} - CNPJ {format_cnpj(cnpj)}")
        apply_subtitle(ws, 2, 1, f"Fechamento Fiscal Detalhado: {mes:02d}/{ano:04d}")
        ws.append([])

        r_idx = 4

        # Separar por tipo
        docs_saida = [d for d in emp_docs if operacao_para(d) == "SAÍDA"]
        docs_entrada = [d for d in emp_docs if operacao_para(d) == "ENTRADA"]

        for secao_nome, docs_secao, fill_cor in [
            ("NOTAS DE SAÍDA (Vendas)", docs_saida, SAIDA_FILL),
            ("NOTAS DE ENTRADA (Compras/Fornecedores)", docs_entrada, ENTRADA_FILL),
        ]:
            if not docs_secao:
                continue

            # Linha de cabeçalho da seção
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=len(headers_item))
            cell_sec = ws.cell(row=r_idx, column=1, value=secao_nome)
            cell_sec.fill = fill_cor
            cell_sec.font = SECTION_FONT
            cell_sec.alignment = openpyxl.styles.Alignment(horizontal="left")
            r_idx += 1

            # Subtotais da seção
            sec_total = sum(float(d.get("valor_total") or 0) for d in docs_secao)
            sec_icms = sum(float(d.get("valor_icms") or 0) for d in docs_secao)
            sec_itens = sum(len(items_por_chave.get(d["chave"], [])) for d in docs_secao)
            ws.cell(row=r_idx, column=1, value=f"{len(docs_secao)} notas | {sec_itens} itens | Total: R$ {sec_total:,.2f} | ICMS: R$ {sec_icms:,.2f}")
            ws.cell(row=r_idx, column=1).font = BOLD_FONT
            r_idx += 1

            # Cabeçalhos
            ws.append(headers_item)
            apply_header_row(ws, r_idx, headers_item)
            r_idx += 1

            # Dados
            for d in docs_secao:
                chave = d.get("chave") or ""
                itens = items_por_chave.get(chave, [])
                emit_cnpj_fmt = format_cnpj(d.get("emitente_cnpj") or "")
                dest_cnpj_fmt = format_cnpj(d.get("destinatario_cnpj") or "")
                dt_fmt = fmt_data(d.get("data_emissao") or "")

                if not itens:
                    ws.append([
                        d.get("numero") or "",
                        dt_fmt,
                        chave,
                        emit_cnpj_fmt,
                        d.get("emitente_nome") or "",
                        dest_cnpj_fmt,
                        d.get("destinatario_nome") or "",
                        d.get("situacao") or "",
                        "", "", "", "XML/Itens não disponíveis", "", "", "",
                        "", 0, 0.0, float(d.get("valor_total") or 0), 0.0,
                    ])
                    for col_i in (18, 19, 20):
                        ws.cell(row=r_idx, column=col_i).number_format = CURRENCY_FMT
                    r_idx += 1
                    continue

                for it in itens:
                    ws.append([
                        d.get("numero") or "",
                        dt_fmt,
                        chave,
                        emit_cnpj_fmt,
                        d.get("emitente_nome") or "",
                        dest_cnpj_fmt,
                        d.get("destinatario_nome") or "",
                        d.get("situacao") or "",
                        it.get("n_item") or "",
                        it.get("codigo") or "",
                        it.get("ean") or "",
                        it.get("descricao") or "",
                        it.get("cfop") or "",
                        it.get("ncm") or "",
                        it.get("cst") or "",
                        it.get("unidade") or "",
                        float(it.get("quantidade") or 0),
                        float(it.get("valor_unitario") or 0),
                        float(it.get("valor_total") or 0),
                        float(it.get("v_icms") or 0),
                    ])
                    for col_i in (17, 18, 19, 20):
                        ws.cell(row=r_idx, column=col_i).number_format = CURRENCY_FMT
                    ws.cell(row=r_idx, column=17).number_format = NUM_FMT
                    r_idx += 1

            # Linha subtotal da seção
            ws.append([
                "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                f"SUBTOTAL {secao_nome.split('(')[0].strip()}",
                sec_itens,
                "",
                sec_total,
                sec_icms,
            ])
            for col_i in range(1, len(headers_item) + 1):
                cell = ws.cell(row=r_idx, column=col_i)
                cell.fill = TOTAL_FILL
                cell.font = BOLD_FONT
                if col_i in (17, 19, 20):
                    cell.number_format = CURRENCY_FMT if col_i != 17 else NUM_FMT
            r_idx += 1

            # Linha em branco entre seções
            r_idx += 1

        # Largura das colunas
        widths = {
            1: 9, 2: 12, 3: 46, 4: 18, 5: 32, 6: 18, 7: 32, 8: 12,
            9: 6, 10: 12, 11: 14, 12: 38, 13: 7, 14: 10, 15: 6,
            16: 8, 17: 11, 18: 14, 19: 16, 20: 14,
        }
        for c_col, w in widths.items():
            ws.column_dimensions[get_column_letter(c_col)].width = w
        ws.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

