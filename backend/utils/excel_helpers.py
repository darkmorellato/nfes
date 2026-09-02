"""Estilos e helpers compartilhados para geração de planilhas Excel (openpyxl)."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1B4F72")
SUB_FONT = Font(name="Calibri", size=10, italic=True, color="555555")
TOTAL_FILL = PatternFill(start_color="EAEDED", end_color="EAEDED", fill_type="solid")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
CURRENCY_FMT = "R$ #,##0.00"
NUM_FMT = "#,##0"
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def format_cnpj(cnpj: str) -> str:
    cnpj = "".join(c for c in str(cnpj or "") if c.isdigit())
    if len(cnpj) == 14:
        return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
    return cnpj


def apply_header_row(ws, row_idx: int, headers: list, fill=None, font=None):
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_title(ws, row: int, col: int, text: str, font=None):
    cell = ws.cell(row, col, text)
    cell.font = font or TITLE_FONT
    return cell


def apply_subtitle(ws, row: int, col: int, text: str, font=None):
    cell = ws.cell(row, col, text)
    cell.font = font or SUB_FONT
    return cell


def auto_adjust_columns(ws, min_width: int = 10, max_width: int = 45):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)
